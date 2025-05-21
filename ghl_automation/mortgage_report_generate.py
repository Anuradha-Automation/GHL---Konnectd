import sys
import psycopg2
import csv
import os
import json
from dotenv import load_dotenv
import openpyxl
from openpyxl.styles import PatternFill, Font

# Load environment variables from .env file
load_dotenv()

# Retrieve environment variables
DB_HOST = os.getenv('DB_HOST')
DB_PORT = os.getenv('DB_PORT', 5432)  # Default PostgreSQL port
DB_NAME = os.getenv('DB_NAME')
DB_USER = os.getenv('DB_USER')
DB_PASSWORD = os.getenv('DB_PASSWORD')
MORTGAGE_LOCATION_IDS = "MORTGAGE_LOCATION_IDS"
# Get logger instance
# Get logger instance (assumed to be a custom logger)
sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), '..')))

from logs_setup_file import setup_logging

module_name = os.path.splitext(os.path.basename(__file__))[0]

logger = setup_logging(module_name)
base_dir = os.path.dirname(os.path.abspath(__file__))

def get_env_list(key: str):
    """Reads an environment variable and splits it into a list if multiple values are present."""
    value = os.getenv(key, "")
    return [v.strip() for v in value.split(",") if v.strip()]


# Function to get the database connection
def get_db_connection():
    return psycopg2.connect(
        host=DB_HOST,
        port=DB_PORT,
        dbname=DB_NAME,
        user=DB_USER,
        password=DB_PASSWORD
    )




# Query to get all contacts with their custom fields
def get_all_contacts_with_custom_fields(location_id):
    query = f"""
    SELECT 
        ccf.contact_id,
        c.first_name,
        c.last_name,
        c.address1,
        c.state,
        jsonb_object_agg(cf.name, ccf.value) AS custom_fields_data
    FROM contact_custom_fields ccf
    JOIN custom_fields cf ON ccf.custom_field_id = cf.id
    JOIN contact c ON ccf.contact_id = c.id
    WHERE c.location_id = '{location_id}'
    GROUP BY ccf.contact_id, c.first_name, c.last_name, c.address1, c.state;
    """
    
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute(query)
    data = cursor.fetchall()
    cursor.close()
    conn.close()
    
    return data

# Write the data to a CSV file
def write_to_csv(data, location):
    print(json.dumps(data, indent=4))
    exit()
    file_path = os.path.join(
        base_dir,
        "xlsm_file_template/Mortgage-Activity-Report-template.xlsx",
    )
    wb1 = openpyxl.load_workbook(file_path, keep_vba=True)
    # Update all sheets
    sheet = wb1["Mortgage Activity Form"]
    start_row = 10  # Starting row for item data
    #print(f"Sample custom fields: {json.dumps(data, indent=4)}")
    row = start_row
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # Yellow color
    green_fill = PatternFill(start_color="00B050", end_color="00B050", fill_type="solid")  # Green
    bold_font = Font(bold=True, size=10) 
    light_green_fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")

    for index, item in enumerate(data):
        if (item[5].get("Current Loan Status Arive", "N/A") and 
            item[5].get('Loan Amount', 0) and 
            item[5].get('Occupancy', "N/A") and 
            item[5].get('Property Type (Housing Type)', "N/A") and 
            item[5].get('Mortgage Type', "N/A") and 
            item[5].get('Loan Purpose', "N/A")):
            sheet[f"A{row}"] = "Not Found Data"
            sheet[f"B{row}"] = item[5].get("Current Loan Status Arive", "N/A")  
            sheet[f"C{row}"] = item[5].get("Loan Funded", "N/A")
            sheet[f"D{row}"] = str(item[1]) + " " + str(item[2])
            sheet[f"E{row}"] = str(item[3])
            sheet[f"F{row}"] = str(item[4])
            sheet[f"G{row}"] = item[5].get('Loan Amount', 0)
            sheet[f"H{row}"] = item[5].get('Occupancy', "N/A") 
            sheet[f"I{row}"] = "Not Found Data"
            sheet[f"J{row}"] = item[5].get('Property Type (Housing Type)', "N/A") 
            sheet[f"K{row}"] = item[5].get('Mortgage Type', "N/A") 
            sheet[f"L{row}"] = item[5].get('Loan Purpose', "N/A") 
            sheet[f"M{row}"] = item[5].get('Compensation', "N/A") 
            sheet[f"N{row}"] = "Not Found Data"
            # ✅ Ensure M column is included in the loop
            for col in ["I", "J", "K", "L", "M"]:
                cell = sheet[f"{col}{row}"]
                cell.fill = light_green_fill  
            # ✅ Apply green fill only to N column
            cell_n = sheet[f"N{row}"]
            cell_n.fill = green_fill  
            row += 1
    static_row = row + 4   
    static_data = {
        "B": ["Status Options:", "Pending", "Not accepted", "Withdrawn", "Denied", "Incomplete", "Funded", ""],
        "C": ["", "File in process", "Approved by lender but not accepted by borrower", 
            "Loan withdrawn by borrower", "Loan denied by lender", 
            "Loan closed for incompleteness", "Loan has funded", ""],
        "H": ["Notes:", "Occupancy", "Owner Occupied", "Second Home", "Investment", "", ""],
        "I": ["", "Position", "First", "Second", "", "", "", ""],
        "J": ["", "Property type", "1-4 Family", "Manufactured", "Multi-Family", "", "", ""],
        "K": ["", "Type", "Conventional", "FHA", "VA", "FSA-RHA", "Reverse", ""],
        "L": ["", "Purpose", "Purchase", "Refi", "Home Imp", "", "", ""],
        "M": ["", "**Total broker check from the lender", "minus any 3rd party fees/reimbursements", "", "", "", "", ""]
    }

    # Transpose dictionary values to align row-wise
    for row_index, values in enumerate(zip(*static_data.values()), start=1):  
        cell_b = sheet[f"B{static_row}"]
        cell_b.value = values[0] if len(values) > 0 else ""
        cell_b.fill = yellow_fill  # Always yellow for column B
        cell_b.font = bold_font  

        sheet[f"C{static_row}"] = values[1] if len(values) > 1 else ""

        # Apply yellow fill to column H for 1st and 2nd row
        cell_h = sheet[f"H{static_row}"]
        cell_h.value = values[2] if len(values) > 2 else ""
        if row_index <= 2:  # First two rows in column H should be yellow
            cell_h.fill = yellow_fill
            cell_h.font = bold_font  

        # Apply yellow fill to column I, J, K, L only for the 2nd row
        for col, idx in zip(["I", "J", "K", "L"], range(3, 7)):
            cell = sheet[f"{col}{static_row}"]
            cell.value = values[idx] if len(values) > idx else ""
            if row_index == 2:  # Only the second row should be yellow
                cell.fill = yellow_fill
                cell.font = bold_font  

        # Apply green fill to column M for row 2 and 3
        if row_index in [2, 3]:  
            # ✅ Apply green fill to M column
            cell_m = sheet[f"M{static_row}"]
            cell_m.value = values[7] if len(values) > 7 else ""  
            cell_m.fill = green_fill  
            cell_m.font = bold_font  

            # ✅ Apply green fill to N column
            cell_n = sheet[f"N{static_row}"]
            cell_n.value = values[8] if len(values) > 8 else ""  
            cell_n.fill = green_fill  
            cell_n.font = bold_font  
        static_row += 1  # Move to the next row
    output_dir = os.path.join(base_dir, "output_files")
    os.makedirs(output_dir, exist_ok=True)
    new_file_path = os.path.join(output_dir, f"updated_invoice_{location}_template.xlsx")
    wb1.save(new_file_path)   

    
    # Log that the CSV file has been generated

# Main function to fetch data and generate the CSV
def main():
    logger.info("Fetching all contacts data with custom fields...")
   
    location_ids = get_env_list(MORTGAGE_LOCATION_IDS)
    for location in location_ids:
        data = get_all_contacts_with_custom_fields(location)
        
        if data:
            logger.info(f"Found {len(data)} records.")
            write_to_csv(data, location)
        else:
            logger.warning("No data found.")

if __name__ == '__main__':
    main()
