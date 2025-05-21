from datetime import datetime
import sys
import time
from utils import  login
import pandas as pd
import openpyxl
from webdriver_configration import driver_confrigration
from selenium.webdriver.common.by import By
import requests
import os
import re
from dotenv import load_dotenv
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.keys import Keys
from openpyxl.styles import PatternFill

# Load environment variables from the .env file
load_dotenv()

# Access the google_credentials path from the .env file
COURSE_PROGRESS_ZAP_WEBHOOK_URL = os.getenv("COURSE_PROGRESS_ZAP_WEBHOOK_URL")
excluded_emails = set(os.getenv("LEADERSHIP_COURSE_EXCLUDED_EMAILS", "").split(","))

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
current_day = datetime.today().strftime("%A")
XLSX_FILE_PATH = os.path.join(BASE_DIR, "progress_report", f"Leadership Leap Course -  {current_day}.xlsx")

def parse_course_data(course_data):
    if course_data:
        match = re.match(r"([^:]+):(.+)\(([^)]+)\)", course_data)
        if match:
            email, course_name, location_id = match.groups()
            return email, course_name.strip(), location_id.strip()
    return None, None, None

# Fetching and parsing course data
course_1_data = os.getenv("CourseProgressReport3")
email_1, course_name_1, location_id_1 = parse_course_data(course_1_data)

sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), '..')))

from logs_setup_file import setup_logging

module_name = os.path.splitext(os.path.basename(__file__))[0]

logger = setup_logging(module_name)


def click_on_all_member_button(driver, max_retries=2):
    """Attempts to click on the 'All Members' button, refreshing if needed."""
    
    for attempt in range(max_retries):
        try:
            body = driver.find_element("tag name", "body")
            for _ in range(8):
                body.send_keys(Keys.ARROW_UP)
                time.sleep(2)
            
            time.sleep(15)
            
            all_members_button = driver.find_element(By.XPATH, "//div[contains(text(), 'All Members')]")
            all_members_button.click()
            logger.info("Successfully clicked on 'All Members' button.")
            time.sleep(5) 
            return  

        except Exception as e:
            logger.warning(f"Attempt {attempt + 1}/{max_retries} failed: {e}")
            if attempt < max_retries - 1:
                logger.info("Refreshing the page and retrying...")
                driver.refresh()
                time.sleep(10) 
            else:
                logger.error("Failed to click 'All Members' after multiple attempts.")


def click_on_next_page(driver):
    next_page_button = driver.find_element(By.XPATH, "//li[@title='Next Page' and not(contains(@class, 'disabled'))]/a")
    next_page_button.click()

def scrape_data_of_user(driver, data, current_row, logger):
    try:
        logger.info(f"Processing row {current_row}")
        time.sleep(15)

        # Click on the user's row
        row_xpath = f"//*[@id='app']/div/div/div/div/div/div/div[2]/div/div[1]/div[2]/table/tbody/tr[{current_row}]/td[1]/div/div/a"
        driver.find_element(By.XPATH, row_xpath).click()
        logger.info(f"Successfully clicked on row {current_row}")
        time.sleep(25)

        # Extract user details
        name = driver.find_element(By.XPATH, "//div[@class='details md:text-lg py-2 w-3/5']//span[@class='font-medium']").text
        email = driver.find_element(By.XPATH, "//div[@class='details md:text-lg py-2 w-3/5']//h6/a").text
        last_login = driver.find_element(By.XPATH, "//div[@class='details md:text-lg py-2 w-3/5']//p[contains(., 'Last Login:')]/span[@class='other-data']").text.strip().lower()

        logger.info(f"Row {current_row} Name: {name}")
        logger.info(f"Row {current_row} Email: {email}")
        logger.info(f"Row {current_row} Last Login: {last_login}")

        module_status = []

        try:
            rows = driver.find_elements(By.XPATH, "//div[contains(@class, 'category-list-items') and "
                                                  "(.//span[contains(text(), 'Module')] or "
                                                  ".//span[contains(text(), 'Introduction to 102 - Managing Others') or "
                                                  "contains(text(), 'Wrap Up') or "
                                                  "contains(text(), 'Welcome to Pillar') or "
                                                  "contains(text(), 'Masterclass')])]")

            for index, row in enumerate(rows, start=1):

                if len(row.find_elements(By.XPATH, ".//img[@alt='Completed']")) > 0:
                    status = "Done"
                elif len(row.find_elements(By.XPATH, ".//span[contains(text(), '%')]")) > 0:
                    percentage = row.find_element(By.XPATH, ".//span[contains(text(), '%')]").text.strip()
                    status = percentage
                elif len(row.find_elements(By.XPATH, ".//i[contains(@class, 'ivu-icon-md-lock')]")) > 0:
                    status = "0%"
                else:
                    status = "No status found"

                module_status.append(status)

        except:
            logger.info("Modules data not found")

        if last_login == "never":
            row_data = [name, email] + [""] * len(module_status)  
        else:
            row_data = [name, email] + module_status  

        data.append(row_data)
        
        logger.info("Successfully retrieved data for all modules")

        time.sleep(5)
        click_on_all_member_button(driver)
        driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
        time.sleep(4)

    except Exception as e:
        logger.info(f"Error processing row {current_row}: {e}")

def page_count_click_next(driver,page_count):
    print("click on next btn according to page btn",page_count)
    for i in range(1,page_count):
        print("click next button times ", i)
        try:
            time.sleep(3)
            click_on_next_page(driver)
        except Exception as e:
            print(e)
            driver.refresh()
            time.sleep(50)
            driver.switch_to.frame("membership-builder")
            time.sleep(7)
            print("refresh page again")
        time.sleep(2)

def extract_data(driver, data, total_page_count, logger):
    for page_count in range(1,total_page_count+1):
        print("page_count------------------------------------", page_count)  
        for current_row in range(1, 11):
            if page_count != 1:
                print(current_row,"current_row",page_count,"page_count")
                page_count_click_next(driver,page_count)
                time.sleep(10)   

            scrape_data_of_user(driver, data,current_row, logger)
        
        print(f"Moving to Page {page_count}")
        logger.info(f"Moving to Page {page_count}")

def save_to_xlsx(data, sheet_name, file_path, logger):
    """Save extracted data into an Excel file with formatting."""
    logger.info(f"Saving data to Excel: {file_path}")
    if not data:
        logger.warning("No data to save.")
        return
    
    headers = ["ND Participants", "Email",
               "Module 1: You’re a Manager…Now What?",
               "Module 2: Internal Foundation",
               "Module 3: Self Awareness & Social Reality",
               "Module 4: Maximizing Communication Using DISC",
               "Module 5: Building Your Personal Brand",
               "Module 6: Prioritization & Time Management",
               "Pillar 102 Intro",
               "Module 1:Situational Leadership",
               "Module 2: Communication",
               "Module 3: Building Your Team Brand",
               "Module 4:Motivation",
               "Module 5:Coaching & Feedback",
               "Module 6:Conflict Management/Sensitive Conversations",
               "Module 7: Strategically Delegation / Accountability",
               "Module 8: Team Prioritization / Team Goal Planning",
               "Wrap Up",
               "Welcome to Pillar",
               "Module 1:Seeing the Big Picture",
               "Module 2: Strategic Goal Setting",
               "Module 3: Building a Strategic Alliance Plan",
               "Module 4: Influence",
               "Module 5: Leadership Presence",
               "Module 6: Business Acumen",
               "Masterclass Wrap Up"
               ] 

    excluded_emails = set(email.lower() for email in os.getenv("LEADERSHIP_COURSE_EXCLUDED_EMAILS", "").split(","))
    filtered_data = [row for row in data if row[headers.index("Email")].lower() not in excluded_emails]
    
    df = pd.DataFrame(filtered_data, columns=headers)


    if os.path.exists(file_path):
        workbook = openpyxl.load_workbook(file_path)
    else:
        workbook = openpyxl.Workbook()

    if "Sheet" in workbook.sheetnames:
        std = workbook["Sheet"]
        workbook.remove(std)

    if sheet_name in workbook.sheetnames:
        workbook.remove(workbook[sheet_name])

    worksheet = workbook.create_sheet(sheet_name)
    
    worksheet.merge_cells(start_row=1, start_column=3, end_row=1, end_column=8)  
    worksheet.merge_cells(start_row=1, start_column=9, end_row=1, end_column=26) 
    
    worksheet.cell(row=1, column=3, value="Pillar 101")
    worksheet.cell(row=1, column=9, value="Pillar 102")
    
    worksheet.append(headers)

    red_fill = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid") 
    blue_fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid") 

    for row_data in filtered_data:
        row_index = worksheet.max_row + 1
        worksheet.append(row_data)

        if row_data[2] == "":
            worksheet[f"A{row_index}"].fill = red_fill

        for col_idx, cell_value in enumerate(row_data[2:], start=3):  
            cell_ref = worksheet.cell(row=row_index, column=col_idx)
            if isinstance(cell_value, str) and cell_value.strip().lower() == "done":
                cell_ref.fill = blue_fill


    os.makedirs(os.path.dirname(file_path), exist_ok=True)

    workbook.save(file_path)
    workbook.close()

    logger.info(f"Excel file '{file_path}' saved successfully!")
    
def send_xlsx_to_zapier(filename, email):
    with open(filename, 'rb') as file:
        files = {'file': (filename, file, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')}
        data = {"action": "xlsx_uploaded", "email": email} 
        response = requests.post(COURSE_PROGRESS_ZAP_WEBHOOK_URL, files=files, data=data)

        if response.status_code == 200:
            logger.info(f"XLSX file {filename} successfully sent to Zapier.")
        else:
            logger.error(f"Failed to send {filename}. Status code: {response.status_code}, Response: {response.text}")       

def read_xlsx_file(email):
    if not os.path.exists(XLSX_FILE_PATH):
        logger.error(f"File not found: {XLSX_FILE_PATH}")
        return
    try:
        df = pd.read_excel(XLSX_FILE_PATH)  
        logger.info(f"Successfully read the XLSX file: {XLSX_FILE_PATH}")
        send_xlsx_to_zapier(XLSX_FILE_PATH, email)
    except Exception as e:
        logger.error(f"Error reading the XLSX file: {str(e)}")

def main_process_of_course(course_name,data):
    
    logger.info("Starting the scrapping process.")
    # Record the start time
    start_time = datetime.now()
    print("Script started at:", start_time)
    logger.info(f"Script started at: {start_time}")
    
    driver = driver_confrigration(logger)
    logger.info("Driver configuration completed.")
    
    login(driver,logger)
    
 
    driver.get(f"https://app.konnectd.io/v2/location/{location_id_1}/memberships/courses/products/")
    time.sleep(50)

    courses_menu = driver.find_element(By.XPATH, "//a[@id='tb_courses']")
    actions = ActionChains(driver)

    actions.move_to_element(courses_menu).perform()
    
    time.sleep(10)

    analytics_option = driver.find_element(By.XPATH, "//a[@id='tb-sub_analytics']")
    analytics_option.click()
    
    logger.info("click on analytics")
    time.sleep(5) 
    driver.refresh()
    time.sleep(20)
    
    try:
        driver.switch_to.frame("membership-builder")
        time.sleep(6)
        logger.info("Switched to iframe 'membership-builder'.")
    except Exception as e:
        logger.error(f"Iframe not found: {e}")
        
    time.sleep(20)
   
    course_progress = driver.find_element(By.XPATH, "//div[contains(@class, 'flex') and contains(@class, 'cursor-pointer') and .//div[contains(text(), 'Course Progress')] and .//div[contains(text(), 'Track progress of your learners')]]")
    course_progress.click()
          
    logger.info("Clicked on course progress:")

    time.sleep(15)
    
    logger.info(f"course name --------{course_name}")
    course_name_element = driver.find_element(By.XPATH, f"//h2[contains(text(), '{course_name}')]")
    course_name_element.click()
    
    logger.info(f"click on {course_name} ")
    
    sheet_according_to_course_name = str(course_name).strip()
    
    time.sleep(20)
    
    last_page_element = driver.find_element(By.XPATH, "//ul[@class='ivu-page']/li[last()-1]/a")
    last_page_number = last_page_element.text
    print("Last Page Number:", last_page_number)
    logger.info(f"Total page count is --{last_page_number}")
    

    total_page_count = int(last_page_number)
    
    extract_data(driver, data, total_page_count, logger)
    
    logger.info(f"all the data of {course_name} succesfully extracted")
    
    save_to_xlsx(data, sheet_according_to_course_name, XLSX_FILE_PATH, logger)
    
    driver.quit()
    time.sleep(5)
          
    logger.info(f"✅ All the data of {course_name} successfully saved in XLSX file! 🎉📊")
    print(f"✅ All the data of {course_name} successfully saved in XLSX file! 🎉📊")
    
data = []    
main_process_of_course(course_name_1,data)
read_xlsx_file(email_1)