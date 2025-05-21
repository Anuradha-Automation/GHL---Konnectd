from datetime import datetime
import sys
import time
from utils import login
import pandas as pd
from webdriver_configration import driver_confrigration
from selenium.webdriver.common.by import By
import requests
import os
import re
from dotenv import load_dotenv
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.keys import Keys
from openpyxl import load_workbook

from logs_setup_file import setup_logging

sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), '..')))

module_name = os.path.splitext(os.path.basename(__file__))[0]

logger = setup_logging(module_name)


is_logged_in = False
# Load environment variables from the .env file
load_dotenv()

# Access the google_credentials path from the .env file

COURSE_PROGRESS_ZAP_WEBHOOK_URL = os.getenv("COURSE_PROGRESS_ZAP_WEBHOOK_URL")
excluded_emails = os.getenv("VONLANE_COURSE_EXCLUDED_EMAILS", "").split(",")

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
current_day = datetime.today().strftime("%A")
XLSX_FILE_PATH = os.path.join(BASE_DIR, "progress_report", f"Vonlane(Bus Drivers and Attendants) -  {current_day}.xlsx")


def parse_course_data(course_data):
    if course_data:
        match = re.match(r"([^:]+):(.+)\(([^)]+)\)", course_data)
        if match:
            email, course_name, location_id = match.groups()
            return email, course_name.strip(), location_id.strip()
    return None, None, None

# Fetching and parsing course data
course_1_data = os.getenv("CourseProgressReport1")
course_2_data = os.getenv("CourseProgressReport2")

email_1, course_name_1, location_id_1 = parse_course_data(course_1_data)
email_2, course_name_2, location_id_2 = parse_course_data(course_2_data)




def click_on_all_member_button(driver, max_retries=3):
    """Attempts to click on the 'All Members' button, refreshing if needed."""
    
    for attempt in range(max_retries):
        try:
            body = driver.find_element("tag name", "body")
            for _ in range(12):
                body.send_keys(Keys.ARROW_UP)
                time.sleep(2)
            
            time.sleep(15)
            
            all_members_button = driver.find_element(By.XPATH, "//div[contains(text(), 'All Members')]")
            all_members_button.click()
            logger.info("Successfully clicked on 'All Members' button.")
            time.sleep(5) 
            return  

        except Exception as e:
            logger.warning(f"Attempt {attempt + 1}/{max_retries} failed: {e}")
            if attempt < max_retries - 1:
                logger.info("Refreshing the page and retrying...")
                driver.refresh()
                time.sleep(10) 
            else:
                logger.error("Failed to click 'All Members' after multiple attempts.")

def click_on_next_page(driver):
    next_page_button = driver.find_element(By.XPATH, "//li[@title='Next Page' and not(contains(@class, 'disabled'))]/a")
    next_page_button.click()

def scrape_data_of_user(driver, data, current_date, current_row, logger):
    try:
        logger.info("start for extracting the user progress")
        time.sleep(15)
     
        logger.info(f"Start for row process {current_row}")
        row_xpath = f"//*[@id='app']/div/div/div/div/div/div/div[2]/div/div[1]/div[2]/table/tbody/tr[{current_row}]/td[1]/div/div/a"
        driver.find_element(By.XPATH, row_xpath).click()
        time.sleep(20)
        logger.info(f"Successfully clicked on row {current_row}")
            
        name = driver.find_element(By.XPATH, "//div[@class='details md:text-lg py-2 w-3/5']//span[@class='font-medium']").text
        email = driver.find_element(By.XPATH, "//div[@class='details md:text-lg py-2 w-3/5']//h6/a").text
        last_login = driver.find_element(By.XPATH, "//div[@class='details md:text-lg py-2 w-3/5']//p[contains(., 'Last Login:')]/span[@class='other-data']").text.strip().lower()

        logger.info(f"Row {current_row} name is :  {name}")
        logger.info(f"Row {current_row} email is :  {email}")
        logger.info(f"Row {current_row} last_login is :  {last_login}")

        try:
            driver.find_element(By.XPATH, "//div[contains(@class, 'post-list-items')]//img[@alt='Completed']")
            welcome_status = "done"
        except:
            welcome_status = "0%"

        if last_login == "never":
            welcome_status = ""
            week_progress = ["  "] * 14 
        else:
            week_row = driver.find_elements(By.XPATH, "//div[contains(@class, 'category-list-items')]//span[contains(text(), 'Week')]")
            week_count = len(week_row)
            logger.info(f"Total Weeks row count is {week_count}")
            time.sleep(5)
            week_progress = []

            for j in range(2, week_count + 1):
                try:
                    time.sleep(2)
                    weeks_row = driver.find_element(By.XPATH, f"//*[@id='app']/div/div/div/div/div/div/div/div[1]/div[2]/div[3]/div/div[{j}]/div")
                    weeks_row.click()
                    time.sleep(7)
                    logger.info(f"Successfully clicked on week row {j}")

                    video_xpath = "//*[@id='app']/div/div/div/div/div/div/div/div[2]/div/div[3]/div/div[1]/span"
                    video = driver.find_element(By.XPATH, video_xpath).text.strip()
                    video_status = "0%" if video == "0%" else "done"
                    quiz_xpath = "//*[@id='app']/div/div/div/div/div/div/div/div[2]/div/div[3]/div/div[2]/span"
                    
                    quiz = driver.find_element(By.XPATH, quiz_xpath).text.strip()
                    quiz_status = "0%" if quiz == "0%" else "done"

                    week_progress.append(video_status)
                    week_progress.append(quiz_status)

                except Exception as e:
                    logger.info(f"Error processing week {j}: {e}")

            while len(week_progress) < 14:
                week_progress.append("  ")

        row_data = [current_date, name, email, welcome_status] + week_progress[:14]
        
        data.append(row_data)

        logger.info("Successfully retrieved data for all weeks")
        time.sleep(5)
        click_on_all_member_button(driver)
        
        driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
        time.sleep(4)
    except Exception as e:
        logger.info(f"Error processing row {current_row}: {e}")

def page_count_click_next(driver,page_count):
    print("click on next btn according to page btn", page_count)
    for i in range(1,page_count):
        print("click next button times ", i)
        try:
            time.sleep(3)
            click_on_next_page(driver)
        except Exception as e:
            print(e)
            driver.refresh()
            time.sleep(50)
            driver.switch_to.frame("membership-builder")
            time.sleep(7)
            print("refresh page again")
        time.sleep(2)

     
def extract_data(driver, data, total_page_count, current_date, logger):
    for page_count in range(1,total_page_count+1):
        for current_row in range(1, 11):
            if page_count != 1:
                print(current_row,"current_row",page_count,"page_count")
                page_count_click_next(driver,page_count)
                time.sleep(10)   

            scrape_data_of_user(driver, data, current_date,current_row, logger)
        
        print(f"Moving to Page {page_count}")
        logger.info(f"Moving to Page {page_count}")
            
def save_to_xlsx(data, sheet_according_to_course_name, logger):
    """Save extracted data based on course name condition."""
    
    logger.info(f"Sheet name based on course name --> {sheet_according_to_course_name}")

    header = ["Date Added", "Name", "Email", "Carter's Welcome/ You're In! Video"]
    for i in range(1, 8):
        header.append(f"Video{i}")
        header.append(f"Quiz{i}")

    # Fill missing columns
    for row in data:
        while len(row) < len(header):
            row.append(" ")

    df = pd.DataFrame(data, columns=header)

    def highlight_welcome(val):
        return 'background-color: #FFCCCB' if isinstance(val, str) and val.strip() == "" else ''

    def highlight_progress(val):
        return 'background-color: #ADD8E6' if val.strip().lower() == "done" else ''

    df_filtered = df[~df["Email"].isin(excluded_emails)]
    
    styled_df = df_filtered.style.map(highlight_welcome, subset=["Carter's Welcome/ You're In! Video"]) \
                        .map(highlight_progress, subset=[f"Video{i}" for i in range(1, 8)] + [f"Quiz{i}" for i in range(1, 8)])

    # Ensure directory exists
    os.makedirs(os.path.dirname(XLSX_FILE_PATH), exist_ok=True)

    if sheet_according_to_course_name == 'Vonlane Bus Drivers: Going the Extra Mile':
        sheet_name = "Bus Drivers"

    elif sheet_according_to_course_name == 'Vonlane Attendants: Going the Extra Mile':
        sheet_name = "Attendants"

    else:
        sheet_name = "Progress Report"
        logger.info(f"Course name '{sheet_according_to_course_name}' does not match. Using default sheet.")

    # Check if the file exists
    if os.path.exists(XLSX_FILE_PATH):
        logger.info(f"Existing report found. Replacing sheet '{sheet_name}'.")
        
        book = load_workbook(XLSX_FILE_PATH)
        
        if sheet_name in book.sheetnames:
            del book[sheet_name]
            book.save(XLSX_FILE_PATH)
            
        with pd.ExcelWriter(XLSX_FILE_PATH, engine="openpyxl", mode="a") as writer:
            styled_df.to_excel(writer, index=False, sheet_name=sheet_name)

        print(f"Sheet '{sheet_name}' replaced successfully in '{XLSX_FILE_PATH}' 🚀")

    else:
        logger.info(f"No existing report found. Creating a new file and adding sheet '{sheet_name}'.")

        with pd.ExcelWriter(XLSX_FILE_PATH, engine="openpyxl") as writer:
            styled_df.to_excel(writer, index=False, sheet_name=sheet_name)

        print(f"Excel file '{XLSX_FILE_PATH}' created successfully with sheet '{sheet_name}' 🚀")

    time.sleep(5)
        
    
def send_xlsx_to_zapier(filename, email):
    with open(filename, 'rb') as file:
        files = {'file': (filename, file, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')}
        data = {"action": "xlsx_uploaded", "email": email}  # Custom parameter
        response = requests.post(COURSE_PROGRESS_ZAP_WEBHOOK_URL, files=files, data=data)

        if response.status_code == 200:
            logger.info(f"XLSX file {filename} successfully sent to Zapier.")
        else:
            logger.error(f"Failed to send {filename}. Status code: {response.status_code}, Response: {response.text}")       

    
def read_xlsx_file(email):
    if os.path.exists(XLSX_FILE_PATH):
        try:
            df = pd.read_excel(XLSX_FILE_PATH)  
            logger.info(f"Successfully read the XLSX file: {XLSX_FILE_PATH}")
            send_xlsx_to_zapier(XLSX_FILE_PATH, email)
        except Exception as e:
            logger.error(f"Error reading the XLSX file: {str(e)}")
    else:
        logger.error(f"File not found: {XLSX_FILE_PATH}")

def main_process_of_course(course_name,data):
    try:
        logger.info("Starting the scrapping process.")
        # Record the start time
        start_time = datetime.now()
        print("Script started at:", start_time)
        logger.info(f"Script started at: {start_time}")
        
        driver = driver_confrigration(logger)
        logger.info("Driver configuration completed.")
        
        login(driver,logger)
        
        driver.get(f"https://app.konnectd.io/v2/location/{location_id_1}/memberships/courses/products/")
        time.sleep(50)

        courses_menu = driver.find_element(By.XPATH, "//a[@id='tb_courses']")
        actions = ActionChains(driver)

        actions.move_to_element(courses_menu).perform()
        
        time.sleep(30)

        analytics_option = driver.find_element(By.XPATH, "//a[@id='tb-sub_analytics']")
        analytics_option.click()
        
        logger.info("click on analytics")
        time.sleep(5) 
        driver.refresh()
        time.sleep(20)
        
        try:
            driver.switch_to.frame("membership-builder")
            time.sleep(6)
            logger.info("Switched to iframe 'membership-builder'.")
        except Exception as e:
            logger.error(f"Iframe not found: {e}")
            
        time.sleep(30)
    
        course_progress = driver.find_element(By.XPATH, "//div[contains(@class, 'flex') and contains(@class, 'cursor-pointer') and .//div[contains(text(), 'Course Progress')] and .//div[contains(text(), 'Track progress of your learners')]]")
        course_progress.click()
            
        logger.info("Clicked on course progress:")

        time.sleep(15)
        
        logger.info(f"course name --------{course_name}")
        course_name_element = driver.find_element(By.XPATH, f"//h2[contains(text(), '{course_name}')]")
        course_name_element.click()
        logger.info(f"click on {course_name} ")
        
        sheet_according_to_course_name = str(course_name).strip()
    
        time.sleep(20)
        
        last_page_element = driver.find_element(By.XPATH, "//ul[@class='ivu-page']/li[last()-1]/a")
        last_page_number = last_page_element.text
        logger.info(f"Total page count is --{last_page_number}")
        
        current_date = datetime.today().strftime("%-m/%d/%y")

        total_page_count = int(last_page_number)

        extract_data(driver, data, total_page_count, current_date,logger)
        
        logger.info(f"✅ All the data of {course_name} successfully saved in XLSX file! 🎉📊")
        
        save_to_xlsx(data, sheet_according_to_course_name, logger)
        
        driver.quit()
        time.sleep(20)             
        return True  
         
    except Exception as e:
        logger.error(f"Error occurred in processing course {course_name}: {e}")
        return False  

        
data = []   
    
success_course_1 = main_process_of_course(course_name_1, data)
success_course_2 = main_process_of_course(course_name_2, data)

if success_course_1 and success_course_2:
    read_xlsx_file(email_1)
